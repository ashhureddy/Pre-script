"""
rrnrbl_checklist.py

Maps the results dict produced by run_validation.run() onto the 63-item
"Legacy - N2e Engineer Checklist" sheet (Checklist_RRNRBL.xlsx) and can
write a filled copy of that exact template (Site ID/FA + Date filled in,
each row's checkbox + Comments column set from the validation results).

Design note on honesty: every row below is wired to a REAL existing check
where one exists (by 'rule' tag - see checks_node.py / checks_sector.py),
a newly-added check where the data was clearly available (EDP field rules,
MME Region, NR_SA tab, FA Code CIQ-vs-RFDS), or left 'manual' when no
reliable signal exists. Nothing here fabricates a pass.
"""
import datetime
import io
import os
import re

import openpyxl

import ciq_edp_reader as cer

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Checklist_RRNRBL.xlsx")

STATUS_META = {
    "match": ("PASS", True),
    "mismatch": ("FAIL", False),
    "manual": ("MANUAL", False),
    "unknown": ("NO DATA", False),
    "na": ("N/A", False),
    "info": ("INFO", False),
}


# ══════════════════════════════════════════════════════════════════════
# Generic aggregation helpers over the existing checks_node/checks_sector
# result lists (every item in those lists already carries a 'status' of
# MATCH / MISMATCH / SKIPPED / INFO - see checks_sector.py).
# ══════════════════════════════════════════════════════════════════════

def _agg(results_list, note_fields=("node", "cell", "note")):
    """Any MISMATCH -> mismatch. Only MATCH/INFO seen -> match. Nothing but
    SKIPPED (or empty) -> unknown (no data to judge, not a pass)."""
    if not results_list:
        return "unknown", "No data (check did not run for this site)."
    real = [r for r in results_list if r.get("status") not in (None, "SKIPPED")]
    bad = [r for r in real if r.get("status") == "MISMATCH"]
    if bad:
        parts = []
        for r in bad[:6]:
            bits = [str(r.get(f)) for f in note_fields if r.get(f)]
            parts.append(": ".join(bits) if bits else str(r))
        more = f" (+{len(bad)-6} more)" if len(bad) > 6 else ""
        return "mismatch", "; ".join(parts) + more
    if real:
        return "match", f"{len(real)} checked, no mismatch."
    skipped_notes = {r.get("note") for r in results_list if r.get("note")}
    return "unknown", "; ".join(sorted(skipped_notes)) or "Skipped for every node (no Pre log / no RFDS)."


def _filter(results_list, rule_prefix):
    return [r for r in results_list if str(r.get("rule", "")).strip() == rule_prefix]


# ══════════════════════════════════════════════════════════════════════
# EDP field-level checks (cabinet naming / port size / port facing /
# bearer VLAN clash / IPv6 bearer+OAM groups). run_validation.py's own
# pipeline never built these - they only exist today in the separate
# HTML tool's EDP Validator - so this ports that exact logic here,
# reading straight off edp_rows (ciq_edp_reader.build_edp_index output).
# ══════════════════════════════════════════════════════════════════════

def _norm(v):
    v = "" if v is None else str(v).strip()
    return "" if v.lower() in ("none", "nan") else v


def _edp_role(edp_row):
    return "PRIMARY" if _norm(edp_row.get("SIAD_PORT_FACING_BBU")) else "SECONDARY"


def _edp_node_rows(edp_rows, node_ids):
    """{node_id: edp_row_or_None} for every node we're checking."""
    out = {}
    for nid in node_ids:
        rows = cer.edp_rows_for_site(edp_rows, nid)
        out[nid] = rows[0] if rows else None
    return out


def _edp_found_status(edp_rows, node_ids):
    rows = _edp_node_rows(edp_rows, node_ids)
    missing = [n for n, r in rows.items() if r is None]
    if not node_ids:
        return "unknown", "No nodes to check."
    if missing:
        return "mismatch", f"Not published in EDP: {', '.join(missing)}"
    return "match", f"{len(node_ids)} node(s) all found in EDP."


def _edp_cabinet_status(edp_rows, node_ids):
    rows = _edp_node_rows(edp_rows, node_ids)
    bad, checked = [], 0
    for nid, r in rows.items():
        if r is None:
            continue
        checked += 1
        cab = _norm(r.get("CABINET"))
        role = _edp_role(r)
        ok = bool(re.match(r"^BBU\s*\d+V?$", cab, re.I)) if cab else False
        if role == "SECONDARY" and cab and not cab.upper().endswith("V"):
            ok = False
        if not ok:
            bad.append(f"{nid}: cabinet '{cab or '(blank)'}' ({role})")
    if not checked:
        return "unknown", "No EDP rows to check."
    if bad:
        return "mismatch", "; ".join(bad[:6])
    return "match", f"{checked} node(s) checked, all pass."


def _edp_port_size_status(edp_rows, node_ids, mm_rows_by_node):
    rows = _edp_node_rows(edp_rows, node_ids)
    bad, checked = [], 0
    for nid, r in rows.items():
        if r is None:
            continue
        mode = _norm(mm_rows_by_node.get(nid, {}).get("BBU Mode")).upper()
        size = _norm(r.get("SIAD_PORT_SIZE_BBU")).upper()
        if not size:
            continue
        checked += 1
        if mode in ("TMBB", "MMBB") and size != "10GE":
            bad.append(f"{nid}: {mode} node shows port size '{size}', expected 10GE")
    if not checked:
        return "unknown", "No SIAD_PORT_SIZE_BBU values to check."
    if bad:
        return "mismatch", "; ".join(bad[:6])
    return "match", f"{checked} node(s) checked, all pass."


def _edp_port_facing_status(edp_rows, node_ids):
    rows = _edp_node_rows(edp_rows, node_ids)
    bad, checked = [], 0
    for nid, r in rows.items():
        if r is None:
            continue
        checked += 1
        role = _edp_role(r)
        facing = _norm(r.get("SIAD_PORT_FACING_BBU"))
        if role == "PRIMARY" and not facing:
            bad.append(f"{nid}: Primary but SIAD_PORT_FACING_BBU is blank")
        if role == "SECONDARY" and facing:
            bad.append(f"{nid}: Secondary but SIAD_PORT_FACING_BBU is populated ('{facing}')")
    if not checked:
        return "unknown", "No EDP rows to check."
    if bad:
        return "mismatch", "; ".join(bad[:6])
    return "match", f"{checked} node(s) checked, all pass."


def _edp_bearer_vlan_status(edp_rows, node_ids):
    rows = _edp_node_rows(edp_rows, node_ids)
    vlans = [(_norm(r.get("BEARER_ENODEB_SB_VLAN_ID")), nid) for nid, r in rows.items() if r]
    vlans = [(v, n) for v, n in vlans if v]
    if not vlans:
        return "unknown", "No BEARER_ENODEB_SB_VLAN_ID values to check."
    seen = {}
    clashes = []
    for v, n in vlans:
        if v in seen and seen[v] != n:
            clashes.append(f"VLAN {v} shared by {seen[v]} and {n}")
        seen[v] = n
    if clashes:
        return "mismatch", "; ".join(clashes[:6])
    return "match", f"{len(vlans)} node(s), no bearer VLAN clash."


def _edp_group_status(edp_rows, node_ids, fields, label):
    rows = _edp_node_rows(edp_rows, node_ids)
    bad, checked = [], 0
    for nid, r in rows.items():
        if r is None:
            continue
        checked += 1
        missing = [f for f in fields if not _norm(r.get(f))]
        if missing:
            bad.append(f"{nid}: missing {', '.join(missing)}")
    if not checked:
        return "unknown", "No EDP rows to check."
    if bad:
        return "mismatch", "; ".join(bad[:6])
    return "match", f"{checked} node(s) checked, all {label} fields present."


IPV6_BEARER_FIELDS = ["IPV6_ENODEB_BEARER_SUBNET_61", "IPV6_ENODEB_SIAD_BEARER_SUB_64",
                       "IPV6_SIAD_BEARER_IP_DEF_ROUTER", "IPV6_ENODEB_BEARER_IP"]
IPV6_OAM_FIELDS = ["OAM_ENODEB_SIAD_OAM_VLAN", "IPV6_ENODEB_OAM_SUBNET_61",
                    "IPV6_ENODEB_SIAD_OAM_SUB_64", "IPV6_SIAD_OAM_IP_DEF_ROUTER", "IPV6_ENODEB_OAM_IP"]


# ══════════════════════════════════════════════════════════════════════
# New checks that had no home anywhere yet: SW-version consistency across
# nodes, MME Region (N2E), NR_SA tab + TAC digit rule, FA Code CIQ-vs-RFDS.
# ══════════════════════════════════════════════════════════════════════

def _sw_consistency_status(sw_version_results):
    versions = {r.get("sw_version") for r in sw_version_results if r.get("sw_version") not in (None, "NOT FOUND")}
    if not versions:
        return "unknown", "No SW version captured from any Pre kget-all log."
    if len(versions) > 1:
        detail = "; ".join(f"{r.get('node')}={r.get('sw_version')}" for r in sw_version_results if r.get("sw_version") not in (None, "NOT FOUND"))
        return "mismatch", f"Mixed SW versions across Pre nodes: {detail}"
    return "match", f"All Pre nodes on {versions.pop()}."


def _mme_region_status(ciq_wb):
    if "Mixed Mode Info" not in ciq_wb.sheetnames:
        return "unknown", "No Mixed Mode Info sheet."
    rows = cer.sheet_rows_as_dicts(ciq_wb["Mixed Mode Info"])
    n2e = [r for r in rows if _norm(r.get("Nokia Site")).lower().startswith("y")]
    if not n2e:
        return "match", "Nokia Site = No for every node — N2E MME Region rule does not apply."
    bad = [r for r in n2e if re.search(r"E-?RAN", _norm(r.get("MME Region")), re.I)
           and not re.search(r"N-?RAN", _norm(r.get("MME Region")), re.I)]
    if bad:
        return "mismatch", "; ".join(f"{_norm(r.get('eNodeB Name'))}: MME Region '{_norm(r.get('MME Region'))}' — should be N-RAN" for r in bad)
    return "match", f"{len(n2e)} N2E node(s), MME Region correctly N-RAN."


def _nr_sa_tac_status(ciq_wb):
    has_nr_sa = "NR_SA" in ciq_wb.sheetnames
    if not has_nr_sa:
        return "na", "No NR_SA tab in this CIQ — SA-carrier TAC rule does not apply."
    if "5G Info" not in ciq_wb.sheetnames:
        return "unknown", "NR_SA tab present but no 5G Info sheet found."
    rows = cer.sheet_rows_as_dicts(ciq_wb["5G Info"])
    bad = []
    checked = 0
    for r in rows:
        nsa_sa = _norm(r.get("NSA/SA")).upper()
        tac = _norm(r.get("nRTAC"))
        cell = _norm(r.get("NRCellDU"))
        if not nsa_sa or not cell:
            continue
        checked += 1
        is_sa = "SA" in nsa_sa and "NSA" not in nsa_sa
        is_nsa = "NSA" in nsa_sa
        if is_sa and len(tac) != 7:
            bad.append(f"{cell}: NSA/SA=SA but nRTAC='{tac}' (expected 7 digits)")
        elif is_nsa and tac not in ("", "0"):
            bad.append(f"{cell}: NSA/SA=NSA but nRTAC='{tac}' (expected blank/0)")
    if not checked:
        return "unknown", "NR_SA tab present but no NSA/SA values read from 5G Info."
    if bad:
        return "mismatch", "; ".join(bad[:6])
    return "match", f"{checked} 5G Info row(s): nRTAC digit-count matches NSA/SA."


def _fa_code_status(site_details, ciq_wb):
    rfds_fa = _norm(site_details.get("fa_code"))
    if "5G Info" not in ciq_wb.sheetnames:
        return "unknown", "No 5G Info sheet (LTE-only build) to compare."
    rows = cer.sheet_rows_as_dicts(ciq_wb["5G Info"])
    ciq_fas = sorted({_norm(r.get("FA Code")) for r in rows if _norm(r.get("FA Code"))})
    if not rfds_fa:
        return "mismatch", "FA Code not found on the RFDS Site Details / 5G Info page."
    if not ciq_fas:
        return "unknown", "No FA Code on the CIQ 5G Info sheet."
    bad = [f for f in ciq_fas if f != rfds_fa]
    if bad:
        return "mismatch", f"RFDS FA Code {rfds_fa} vs CIQ FA Code(s) {', '.join(bad)}"
    return "match", f"RFDS and CIQ FA Code both {rfds_fa}."


def _xmu_vs_rfds_status(enb_rows_all, node_ids, rfds_pages):
    has_xmu_nodes = []
    for nid in node_ids:
        row = cer.find_enb_row(enb_rows_all, nid)
        if not row:
            continue
        x1, x2 = _norm(row.get("1st XMU")).upper(), _norm(row.get("2nd XMU")).upper()
        if x1 not in ("", "NO", "N/A", "NOT USED") or x2 not in ("", "NO", "N/A", "NOT USED"):
            has_xmu_nodes.append(nid)
    if not has_xmu_nodes:
        return "match", "No node shows a 1st/2nd XMU in CIQ eNB Info."
    if not rfds_pages:
        return "unknown", f"{', '.join(has_xmu_nodes)} show XMU in CIQ, but no RFDS PDF was provided to check."
    full_text = " ".join(rfds_pages.values()).upper() if isinstance(rfds_pages, dict) else ""
    if "XMU" not in full_text:
        return "mismatch", f"{', '.join(has_xmu_nodes)} show XMU in CIQ eNB Info, but 'XMU' does not appear anywhere in the RFDS PDF text."
    return "match", f"{len(has_xmu_nodes)} node(s) with XMU in CIQ — RFDS PDF text also mentions XMU."


# ══════════════════════════════════════════════════════════════════════
# The 63-row checklist definition. `row` = exact Excel row in
# Checklist_RRNRBL.xlsx ("Legacy - N2e Engineer Checklist" sheet).
# `check` is a zero-arg callable returning (status, detail), or None
# for a manual item.
# ══════════════════════════════════════════════════════════════════════

def _edp_controller_status(edp_rows, controller_ids):
    """NEW - Controller/ANCEQ checks (cabinet naming/port-size/etc. above are
    Primary/Secondary-only). controller_ids: list of EDP SITE_NAME values for
    Controller rows (= the CIQ's Controller Info 'Controller ID').
    IPv6 ANCEQ fields are checked as informational only - a real, valid EDP
    row was found with every IPv6 ANCEQ_* field genuinely blank, so treating
    it as a required field would be a false mismatch."""
    if not controller_ids:
        return "na", "No Controller node in this CIQ's Controller Info sheet."
    bad, checked = [], 0
    ipv6_present = 0
    for cid in controller_ids:
        rows = cer.edp_rows_for_site(edp_rows, cid)
        if not rows:
            bad.append(f"{cid}: not published in EDP")
            continue
        r = rows[0]
        checked += 1
        missing = [f for f in ("ANCEQ_TYPE", "ANCEQ_NAME", "ANCEQ_SIAD_IP_HOST_1", "ANCEQ_SIAD_IP_HOST_2") if not _norm(r.get(f))]
        if missing:
            bad.append(f"{cid}: missing {', '.join(missing)}")
        if _norm(r.get("ANCEQ_SIAD_IPV6_HOST_1")):
            ipv6_present += 1
    if not checked:
        return "unknown", "; ".join(bad) if bad else "No EDP rows to check."
    if bad:
        return "mismatch", "; ".join(bad[:6])
    return "match", f"{checked} controller(s) checked (IPv4 required fields all present; {ipv6_present} also have IPv6)."


def _edp_ptp_status(edp_rows, node_ids):
    """NEW - EDP's own PTP fields (SIAD_PTP_VLAN_ID + the PTP_VLAN_SUBNET_30 /
    PTP_SIAD_INTERFACE_IP / PTP_CAB_INTERFACE_IP group), confirmed present on
    a real published EDP row. Separate from the kget-log-side PTP guess in
    pre_extract.extract_ptp_status() - this one reads data this backend
    definitely has."""
    rows = _edp_node_rows(edp_rows, node_ids)
    bad, checked, no_ptp = [], 0, 0
    for nid, r in rows.items():
        if r is None:
            continue
        vlan = _norm(r.get("SIAD_PTP_VLAN_ID"))
        if not vlan:
            no_ptp += 1
            continue
        checked += 1
        missing = [f for f in ("PTP_VLAN_SUBNET_30", "PTP_SIAD_INTERFACE_IP", "PTP_CAB_INTERFACE_IP") if not _norm(r.get(f))]
        if missing:
            bad.append(f"{nid}: PTP VLAN {vlan} set but missing {', '.join(missing)}")
    if bad:
        return "mismatch", "; ".join(bad[:6])
    if checked:
        return "match", f"{checked} node(s) with PTP configured, all required fields present."
    return "info", f"No node declares a PTP VLAN in EDP ({no_ptp} checked) — PTP may not be in scope for this build."


def build_checklist(results, site_details, ciq_wb, edp_rows, node_ids, rfds_pages=None):
    mm_rows = cer.mixed_mode_rows(ciq_wb) if ciq_wb else []
    mm_by_node = {}
    for r in mm_rows:
        n = _norm(r.get("Node to be built as")) or _norm(r.get("eNodeB Name"))
        if n:
            mm_by_node[n] = r
    enb_rows_all = cer.enb_info_rows(ciq_wb) if ciq_wb else []

    board_type = results.get("board_type", [])
    identity = results.get("identity", [])

    def edp_field(fields, label):
        return lambda: _edp_group_status(edp_rows, node_ids, fields, label)

    rows = [
        (13, "Major showstopper check", None, "SW should be match with ENM", "NR/Radio",
         lambda: _sw_consistency_status(results.get("sw_version", []))),

        (15, "EDP check", "EDP vs Site", "site_name", "NR/Radio", lambda: _edp_found_status(edp_rows, node_ids)),
        (16, "EDP check", "EDP vs Site", "cabinet", "Radio", lambda: _edp_cabinet_status(edp_rows, node_ids)),
        (17, "EDP check", "EDP vs Site", "bbu_type", "Radio", lambda: _agg(board_type)),
        (18, "EDP check", "EDP vs Site", "node_model", "Radio", lambda: _agg(board_type)),
        (19, "EDP check", "EDP vs Site", "siad_port_size_bbu", "Radio", lambda: _edp_port_size_status(edp_rows, node_ids, mm_by_node)),
        (20, "EDP check", "EDP vs Site", "siad_port_facing_bbu", "Radio", lambda: _edp_port_facing_status(edp_rows, node_ids)),
        (21, "EDP check", "EDP vs Site", "bearer_enodeb_sb_vlan_id", "Radio", lambda: _edp_bearer_vlan_status(edp_rows, node_ids)),
        (22, "EDP check", "EDP vs Site", "ipv6_siad_bearer_ip_def_router", "Radio", edp_field(IPV6_BEARER_FIELDS, "IPv6 bearer")),
        (23, "EDP check", "EDP vs Site", "ipv6_enodeb_bearer_ip", "Radio", edp_field(IPV6_BEARER_FIELDS, "IPv6 bearer")),
        (24, "EDP check", "EDP vs Site", "oam_enodeb_siad_oam_vlan", "Radio", edp_field(IPV6_OAM_FIELDS, "IPv6 OAM")),
        (25, "EDP check", "EDP vs Site", "ipv6_siad_oam_ip_def_router", "Radio", edp_field(IPV6_OAM_FIELDS, "IPv6 OAM")),
        (26, "EDP check", "EDP vs Site", "ipv6_enodeb_oam_ip", "Radio", edp_field(IPV6_OAM_FIELDS, "IPv6 OAM")),

        (28, "RFDS Checks", "Pre Vs RFDS Sheet in QWEST", "FACode", "Radio", lambda: _fa_code_status(site_details, ciq_wb)),
        (29, "RFDS Checks", None, "JobDetail", "Radio", None),
        (30, "RFDS Checks", None, "NonRFInventoryDetails(Final)", "Radio", lambda: _agg(board_type)),
        (31, "RFDS Checks", None, "CellDetails(Final) -- CellID / RCN /RRH", "Radio", lambda: _agg(results.get("cells_vs_rfds", []))),
        (32, "RFDS Checks", None, "AntennaPositionDetails", "Radio", None),
        (33, "RFDS Checks", None, "Plumbing Diagram (TMA)", "Radio", None),

        (35, "CIQ tabs checks", "Revision History", "All Confirmation checks", "NR/Radio", None),
        (36, "CIQ tabs checks", "Mixed Mode Info Tab", "eNBId and gNBId ENM vs CIQ", "NR/Radio", lambda: _agg(identity)),
        (37, "CIQ tabs checks", "Mixed Mode Info Tab", "MME Region", "NR/Radio", lambda: _mme_region_status(ciq_wb)),
        (38, "CIQ tabs checks", "Mixed Mode Info Tab", "Primary & secondary node matches RFDS", "Radio", lambda: _agg(results.get("primary_secondary", []))),

        (39, "CIQ tabs checks", "5g info", "NRCellDU/NRCellCU ENM vs CIQ", "NR/Radio", lambda: _agg(results.get("cells_vs_rfds", []))),
        (40, "CIQ tabs checks", "5g info", "nRTAC/cellLocalId ENM Vs CIQ", "NR/Radio", lambda: _agg(results.get("cell_id_vs_rfds", []))),
        (41, "CIQ tabs checks", "5g info", "arfcnDL/arfcnUL/bSChannelBwDL ENM Vs CIQ", "NR/Radio", lambda: _agg(results.get("params_5g", []))),
        (42, "CIQ tabs checks", "5g info", "RBB Type vs no.ofrx/tx from ENM", "Radio", lambda: _agg(results.get("params_5g", []))),
        (43, "CIQ tabs checks", "5g info", "DSS check", "NR/Radio", None),
        (44, "CIQ tabs checks", "5g info", "ssbFrequency/ssbOffset/ssbDuration", "NR/Radio", lambda: _agg(results.get("params_5g", []))),
        (45, "CIQ tabs checks", "5g info", "NSA/SA", "NR/Radio", lambda: _agg(results.get("nr_tac", []))),
        (46, "CIQ tabs checks", "5g info", "BBU Type should match with RFDS and CIQ", "NR/Radio", lambda: _agg(board_type)),
        (47, "CIQ tabs checks", "5g info", "Cell/RRU/Beam/Antenna/Tilt must same as RFDS", "Radio", lambda: _agg(results.get("cells_vs_rfds", []))),
        (48, "CIQ tabs checks", "5g info", "NR TAC - Existing sectors", "NR/Radio", lambda: _agg(results.get("nr_tac", []))),
        (49, "CIQ tabs checks", "5g info", "NR TAC - newly added Carriers - NSA=0 & SA=7 digit", "NR/Radio", lambda: _nr_sa_tac_status(ciq_wb)),
        (50, "CIQ tabs checks", "5g info", "6472/AIR-6449/AIR6419 - SEF/FRU", "Radio", lambda: _agg(results.get("sef_fru", []))),
        (51, "CIQ tabs checks", "5g info", "Unique Port for 5G/LTE separate radio", "Radio", lambda: _agg(results.get("port_uniqueness", []))),

        (52, "CIQ tabs checks", "gNB Info", "gNBId/gNodeB Name matches Mixed Mode Info", "NR/Radio", lambda: _agg(identity)),
        (53, "CIQ tabs checks", "gNB Info", "DU type same as 5G Info tab", "NR/Radio", lambda: _agg(board_type)),

        (54, "CIQ tabs checks", "eNB Info", "eNBId/eNodeB Name matches Mixed Mode Info", "NR/Radio", lambda: _agg(identity)),
        (55, "CIQ tabs checks", "eNB Info", "BBU Type should match with RFDS", "Radio", lambda: _agg(board_type)),
        (56, "CIQ tabs checks", "eNB Info", "TAC Value", "NR/Radio", lambda: _agg(results.get("tac", []))),

        (57, "CIQ tabs checks", "eUtran Parameters Tab", "earfcnDl/dlChannelBandwidth ENM vs CIQ", "NR/Radio", lambda: _agg(results.get("params_4g", []))),
        (58, "CIQ tabs checks", "eUtran Parameters Tab", "RBB type/noOfTx/noOfRx - ISDLONLY", "NR/Radio", lambda: _agg(results.get("params_4g", []))),
        (59, "CIQ tabs checks", "eUtran Parameters Tab", "cellId ENM vs CIQ (SOW)", "NR/Radio", lambda: _agg(results.get("cell_id_vs_rfds", []))),
        (60, "CIQ tabs checks", "eUtran Parameters Tab", "EutranCellFDDId/beamDirection vs RFDS", "Radio", None),
        (61, "CIQ tabs checks", "eUtran Parameters Tab", "electricalAntennaTilt integer", "Radio", lambda: _agg(results.get("params_4g", []))),
        (62, "CIQ tabs checks", "eUtran Parameters Tab", "configuredOutputPower depends on RRU type", "Radio", None),
        (63, "CIQ tabs checks", "eUtran Parameters Tab", "TxRx/RBB Type vs Single/Double RILink", "Radio", lambda: _agg(results.get("params_4g", []))),
        (64, "CIQ tabs checks", "eUtran Parameters Tab", "Compare Sectorid With Carrier Progression", "Radio", None),
        (65, "CIQ tabs checks", "eUtran Parameters Tab", "PCI uniqueness", "Radio", lambda: _agg(results.get("pci_4g", []) + results.get("pci_5g", []))),
        (66, "CIQ tabs checks", "eUtran Parameters Tab", "Pre-existing node cellId vs ENM & RFDS", "NR/Radio", lambda: _agg(results.get("cell_id_vs_rfds", []))),
        (67, "CIQ tabs checks", "eUtran Parameters Tab", "Riport should be unique", "Radio", lambda: _agg(results.get("xmu_port_overlap", []))),
        (68, "CIQ tabs checks", "eUtran Parameters Tab", "tmaType/tmaConfiguration", "Radio", None),
        (69, "CIQ tabs checks", "eUtran Parameters Tab", "antenna model", "Radio", lambda: _agg(results.get("cells_vs_rfds", []))),
        (70, "CIQ tabs checks", "eUtran Parameters Tab", "XMU Validation vs RFDS", "Radio", lambda: _xmu_vs_rfds_status(enb_rows_all, node_ids, rfds_pages)),
        (71, "CIQ tabs checks", "eUtran Parameters Tab", "ENM Validation - site locator (B2E)", "Radio", None),

        (72, "CIQ tabs checks", "Losses and delay", "Losses/delay matches FDD and TxRx", "Radio", None),
        (73, "CIQ tabs checks", "Antenna Information", "AntennaUnit/AntennaSubunit unique band-wise", "Radio", lambda: _agg(results.get("antenna", []))),
        (74, "CIQ tabs checks", "Sector Movement / Deletion sheet", "Source/target cells match ENM/eUtran", "NR/Radio", lambda: _agg(results.get("cell_id_vs_rfds", []))),

        # Rows 75-76 are new in the updated template (they pushed the old
        # "Pre checks" block from 75-79 down to 77-81). Both are EDP/ENM IP
        # comparisons this project has no automated check for, so they're
        # manual rather than silently reusing an unrelated check's result.
        (75, "IP Validation Pre Vs EDP", None, "NodeB bearer IP / VLAN ID / router default IP vs EDP (board swap node)", "Radio", None),
        (76, "Rehoming sites ( IP Verification )", None, "Existing IP/VLAN of all nodes: EDP vs ENM (Daffi node rehoming)", "Radio", None),

        (78, "Pre checks", "ENM Pre-checks", "Radio Ports", "Radio", lambda: _agg(results.get("radio_type", []))),
        (79, "Pre checks", "ENM Pre-checks", "RfBranch", "Radio", None),
        (80, "Pre checks", "ENM Pre-checks", "Sharing Radio", "Radio", lambda: _agg(results.get("radio_sharing", []))),
        (81, "Pre checks", "ENM Pre-checks", "SSNALIST", "Radio", None),
    ]

    out = []
    for row, cat, sub, item, tag, check in rows:
        if check is None:
            status, detail = "manual", "No automated check exists for this item."
        else:
            try:
                status, detail = check()
            except Exception as e:  # never let one bad check take down the whole checklist
                status, detail = "unknown", f"Check raised an error: {e}"
        out.append({"row": row, "cat": cat, "sub": sub, "item": item, "tag": tag,
                    "status": status, "detail": detail})
    return out


# ══════════════════════════════════════════════════════════════════════
# Fill the real template.
# ══════════════════════════════════════════════════════════════════════

_FPB_PART = "xl/featurePropertyBag/featurePropertyBag.xml"
_FPB_CONTENT_TYPE = "application/vnd.ms-excel.featurepropertybag+xml"
_FPB_REL_TYPE = "http://schemas.microsoft.com/office/2022/11/relationships/FeaturePropertyBag"


def _restore_native_checkboxes(filled_bytes, template_path):
    """openpyxl's save() silently drops xl/featurePropertyBag/featurePropertyBag.xml
    - the part that marks C-column cells as Excel's native interactive
    Checkbox control (confirmed by a real load->set value->save round-trip:
    the part vanishes even though the underlying boolean cell value is
    preserved). Without it, Excel still shows the right TRUE/FALSE value but
    the checkbox widget itself is gone. This copies that part (and its two
    small registration entries) from the original template's zip into the
    filled workbook's zip after openpyxl is done, so the checkboxes stay
    exactly as clickable as they were in the template you uploaded."""
    import zipfile

    with zipfile.ZipFile(template_path) as tz:
        if _FPB_PART not in tz.namelist():
            return filled_bytes  # template has no native checkboxes to restore
        fpb_xml = tz.read(_FPB_PART)

    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(filled_bytes)) as src, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "[Content_Types].xml":
                text = data.decode("utf-8")
                if _FPB_PART.split("xl/")[1] not in text and "featurePropertyBag" not in text:
                    text = text.replace(
                        "</Types>",
                        f'<Override PartName="/{_FPB_PART}" ContentType="{_FPB_CONTENT_TYPE}"/></Types>',
                    )
                data = text.encode("utf-8")
            elif item.filename == "xl/_rels/workbook.xml.rels":
                text = data.decode("utf-8")
                if _FPB_REL_TYPE not in text:
                    existing_ids = [int(rid) for rid in re.findall(r'Id="rId(\d+)"', text)]
                    new_id = f"rId{max(existing_ids, default=0) + 1}"
                    text = text.replace(
                        "</Relationships>",
                        f'<Relationship Id="{new_id}" Type="{_FPB_REL_TYPE}" '
                        f'Target="featurePropertyBag/featurePropertyBag.xml"/></Relationships>',
                    )
                data = text.encode("utf-8")
            dst.writestr(item, data)
        if _FPB_PART not in src.namelist():
            dst.writestr(_FPB_PART, fpb_xml)
    out.seek(0)
    return out.read()


def fill_checklist_xlsx(checklist, site_id_fa, engineer_name=None, sow=None, date_str=None,
                         template_path=TEMPLATE_PATH, manual_overrides=None):
    """manual_overrides: optional {row_number: {'done': bool, 'comment': str}}
    for rows whose status is 'manual' - lets a person's own checkbox/comment
    (entered in the Streamlit UI) override the generic 'no automated check'
    placeholder text before this gets written out."""
    manual_overrides = manual_overrides or {}
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Legacy - N2e Engineer Checklist"]

    ws["B8"] = site_id_fa or ""
    ws["B9"] = date_str or datetime.date.today().strftime("%m/%d/%Y")
    if engineer_name:
        ws["B7"] = engineer_name
    if sow:
        ws["B10"] = sow

    for entry in checklist:
        r = entry["row"]
        override = manual_overrides.get(r)
        if entry["status"] == "manual" and override is not None:
            ws[f"C{r}"] = bool(override.get("done"))
            comment = (override.get("comment") or "").strip()
            ws[f"E{r}"] = f"[MANUAL — user-confirmed] {comment}" if comment else "[MANUAL — marked done, no comment]" if override.get("done") else "[MANUAL] Not yet reviewed."
            continue
        ws[f"C{r}"] = (entry["status"] == "match")
        label, _ = STATUS_META.get(entry["status"], ("", False))
        comment = entry["detail"] or ""
        ws[f"E{r}"] = f"[{label}] {comment}" if label else comment

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _restore_native_checkboxes(buf.read(), template_path)
