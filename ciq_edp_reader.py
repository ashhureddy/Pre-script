"""
CIQ (Excel) and EDP (legacy .xls) readers for the fields the node-level
Pre-checks-validation rules need. Kept separate from pre_extract.py since
these read spreadsheets rather than the kget-all log text.
"""
import openpyxl
import xlrd


def sheet_rows_as_dicts(ws):
    """First row = header. Returns a list of dicts, one per data row."""
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    header = [str(h).strip() if h is not None else '' for h in header]
    out = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        out.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    return out


def load_ciq(path):
    """Open a CIQ workbook. Falls back to a stripped copy when openpyxl
    rejects the stylesheet.

    Confirmed on a real CIQ: some files carry 6-digit RGB colour values (or
    a bare '0') where the spec requires 8-digit aRGB, and openpyxl refuses to
    open the whole workbook over it - 'Colors must be aRGB hex values'. This
    tool only ever reads cell values, never formatting, so the fix is to pad
    those values to valid aRGB in an in-memory copy and open that. The
    original file is never modified."""
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except ValueError as exc:
        if 'aRGB' not in str(exc) and 'stylesheet' not in str(exc):
            raise
        return _load_ciq_with_repaired_styles(path)


def _load_ciq_with_repaired_styles(path):
    import io
    import re
    import shutil
    import zipfile

    def _fix(match):
        val = match.group(1)
        if re.fullmatch(r'[0-9A-Fa-f]{8}', val):
            return match.group(0)
        if re.fullmatch(r'[0-9A-Fa-f]{6}', val):
            return f'rgb="FF{val}"'
        return 'rgb="FF000000"'

    buf = io.BytesIO()
    with zipfile.ZipFile(path) as src, zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == 'xl/styles.xml':
                text = data.decode('utf-8', 'replace')
                data = re.sub(r'rgb="([^"]*)"', _fix, text).encode('utf-8')
            dst.writestr(item, data)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True)


def mixed_mode_rows(ciq_wb):
    return sheet_rows_as_dicts(ciq_wb['Mixed Mode Info'])


def enb_info_rows(ciq_wb):
    return sheet_rows_as_dicts(ciq_wb['eNB Info'])


def gnb_info_5g_info_rows(ciq_wb):
    return sheet_rows_as_dicts(ciq_wb['5G Info'])


def find_mm_row(mm_rows, node_id):
    """Find a Mixed Mode Info row where the node is Primary (Node to be built
    as) OR Secondary (eNodeB Name / gNodeB Name), matched case-insensitively."""
    nid = str(node_id).strip().upper()
    for row in mm_rows:
        candidates = (row.get('Node to be built as'), row.get('eNodeB Name'), row.get('gNodeB Name'))
        if any(str(c).strip().upper() == nid for c in candidates if c is not None):
            return row
    return None


def find_enb_row(enb_rows, node_id):
    nid = str(node_id).strip().upper()
    for row in enb_rows:
        if str(row.get('eNodeB Name', '')).strip().upper() == nid:
            return row
    return None


# ---------------------------------------------------------------------------
# EDP (.xls) — locates the real header row dynamically (row 25 in the samples,
# but not guaranteed fixed — matches QUICKIX's own locate_edp_header_row
# philosophy of not trusting a hardcoded row number).
# ---------------------------------------------------------------------------

def load_edp(path):
    wb = xlrd.open_workbook(path)
    return wb.sheet_by_index(0)


def locate_edp_header_row(ws):
    for r in range(ws.nrows):
        first_cell = ws.cell_value(r, 0)
        if str(first_cell).strip() == 'EDP_SITE_ID':
            return r
    raise ValueError("Could not locate EDP header row (expected 'EDP_SITE_ID' in column A)")


def build_edp_index(ws):
    """Returns (header_list, rows) where rows is a list of dicts keyed by
    header name, one per EDP data row (there can be several rows per site,
    e.g. one per SIAD port entry)."""
    header_row = locate_edp_header_row(ws)
    header = [str(ws.cell_value(header_row, c)).strip() for c in range(ws.ncols)]
    rows = []
    for r in range(header_row + 1, ws.nrows):
        rows.append({header[c]: ws.cell_value(r, c) for c in range(ws.ncols)})
    return header, rows


def edp_rows_for_site(edp_rows, node_id):
    nid = str(node_id).strip().upper()
    return [r for r in edp_rows if str(r.get('SITE_NAME', '')).strip().upper() == nid]


def edp_primary_secondary(edp_rows, node_ids):
    """Rule #3/#31: for each node id in node_ids, look up its EDP row(s) and
    determine Primary (SIAD_PORT_FACING_BBU populated) vs Secondary (blank).
    Returns {node_id: 'PRIMARY'|'SECONDARY'|'NOT FOUND IN EDP'}."""
    result = {}
    for nid in node_ids:
        rows = edp_rows_for_site(edp_rows, nid)
        if not rows:
            result[nid] = 'NOT FOUND IN EDP'
            continue
        port = rows[0].get('SIAD_PORT_FACING_BBU')
        result[nid] = 'PRIMARY' if str(port).strip() not in ('', 'None') else 'SECONDARY'
    return result


def find_revision_history_sheet(ciq_wb):
    """Same fuzzy match as QUICKIX's findRevisionHistorySheet(): exact name
    'Revision History' (any case/whitespace) first, else any sheet whose
    name contains 'revision'."""
    for name in ciq_wb.sheetnames:
        if name.strip().lower() == 'revision history':
            return name
    for name in ciq_wb.sheetnames:
        if 'revision' in name.lower():
            return name
    return None


def read_revision_history(ciq_wb):
    """Pulls columns A-E as a raw grid (the sheet mixes two stacked mini-
    tables, so no single header row is forced) — same approach as QUICKIX's
    extractRevisionHistory(). Returns (sheet_name, rows) where rows is a
    list of 5-value lists, blank rows dropped; (None, []) if no such sheet."""
    sheet_name = find_revision_history_sheet(ciq_wb)
    if not sheet_name:
        return None, []
    ws = ciq_wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        five = [row[i] if i < len(row) and row[i] is not None else '' for i in range(5)]
        if any(str(v).strip() != '' for v in five):
            rows.append(five)
    return sheet_name, rows
